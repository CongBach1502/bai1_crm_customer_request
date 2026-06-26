from odoo import models, fields
import base64
import io

from openpyxl import load_workbook
class ImportRequestWizard(models.TransientModel):
    _name = "crm.request.import.wizard"
    _description = "Import Request Wizard"

    file = fields.Binary(
        string="Excel File",
        required=True
    )
    lead_id = fields.Many2one("crm.lead")

    file_name = fields.Char(
        string="File Name"
    )
    def action_import(self):

     data = base64.b64decode(self.file)

     workbook = load_workbook(io.BytesIO(data))

     sheet = workbook.active

     for row in sheet.iter_rows(min_row=2, values_only=True):

        product_name = row[0]
        qty = row[1]
        date = row[2]
        description = row[3]

        product = self.env["product.product"].search([
            ("name", "=", product_name)
        ], limit=1)

        if product:

            self.env["crm.customer.request"].create({
                "opportunity_id": self.lead_id.id,
                "product_id": product.id,
                "qty": qty,
                "date": date,
                "description": description,
            })

     return {
        "type": "ir.actions.act_window_close",
    }